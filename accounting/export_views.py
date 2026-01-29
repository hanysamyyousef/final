import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .reports import AccountingReports
from django.utils.translation import gettext as _

@login_required
def export_trial_balance_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    cost_center_id = request.GET.get('cost_center')
    
    report = AccountingReports.get_trial_balance(start_date, end_date, cost_center_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _("ميزان المراجعة")
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = True
    ws.sheet_view.rightToLeft = True

    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = _("ميزان المراجعة")
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    if start_date or end_date:
        ws.merge_cells('A2:F2')
        ws['A2'] = f"{_('من')} {start_date or '...'} {_('إلى')} {end_date or '...'}"
        ws['A2'].alignment = Alignment(horizontal='center')

    # Table Headers
    headers = [
        _("كود الحساب"), _("اسم الحساب"), 
        _("مدين (مجاميع)"), _("دائن (مجاميع)"),
        _("مدين (أرصدة)"), _("دائن (أرصدة)")
    ]
    
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    header_font = Font(bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    row_num = 5
    for row_data in report['data']:
        ws.cell(row=row_num, column=1).value = row_data['account'].code
        ws.cell(row=row_num, column=2).value = row_data['account'].name
        ws.cell(row=row_num, column=3).value = float(row_data['debit'])
        ws.cell(row=row_num, column=4).value = float(row_data['credit'])
        ws.cell(row=row_num, column=5).value = float(row_data['balance_debit']) if row_data['balance_debit'] > 0 else 0
        ws.cell(row=row_num, column=6).value = float(row_data['balance_credit']) if row_data['balance_credit'] > 0 else 0
        row_num += 1

    # Footer
    ws.cell(row=row_num, column=1).value = _("الإجمالي")
    ws.cell(row=row_num, column=1).font = Font(bold=True)
    ws.cell(row=row_num, column=5).value = float(report['total_debit'])
    ws.cell(row=row_num, column=6).value = float(report['total_credit'])
    ws.cell(row=row_num, column=5).font = Font(bold=True)
    ws.cell(row=row_num, column=6).font = Font(bold=True)

    # Styling
    for row in ws.iter_rows(min_row=4, max_row=row_num):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=trial_balance.xlsx'
    wb.save(response)
    return response

@login_required
def export_profit_loss_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    cost_center_id = request.GET.get('cost_center')
    
    report = AccountingReports.get_profit_loss(start_date, end_date, cost_center_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _("قائمة الدخل")
    ws.sheet_view.rightToLeft = True

    ws.merge_cells('A1:B1')
    ws['A1'] = _("قائمة الدخل")
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Income Section
    ws.append([])
    ws.append([_("الإيرادات")])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color='008000')
    
    for item in report['income']:
        ws.append([item['account'].name, float(item['balance'])])
    
    ws.append([_("إجمالي الإيرادات"), float(report['total_income'])])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

    # Expenses Section
    ws.append([])
    ws.append([_("المصروفات")])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color='FF0000')
    
    for item in report['expense']:
        ws.append([item['account'].name, float(item['balance'])])
    
    ws.append([_("إجمالي المصروفات"), float(report['total_expense'])])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

    # Net Profit
    ws.append([])
    ws.append([_("صافي الربح / الخسارة"), float(report['net_profit'])])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True, size=12)
    if report['net_profit'] >= 0:
        ws.cell(row=ws.max_row, column=2).font = Font(bold=True, color='008000')
    else:
        ws.cell(row=ws.max_row, column=2).font = Font(bold=True, color='FF0000')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=profit_loss.xlsx'
    wb.save(response)
    return response

@login_required
def export_balance_sheet_excel(request):
    date = request.GET.get('date')
    cost_center_id = request.GET.get('cost_center')
    
    report = AccountingReports.get_balance_sheet(date, cost_center_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _("الميزانية العمومية")
    ws.sheet_view.rightToLeft = True

    ws.merge_cells('A1:B1')
    ws['A1'] = _("الميزانية العمومية")
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Assets
    ws.append([])
    ws.append([_("الأصول")])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color='0000FF')
    for item in report['assets']:
        ws.append([item['account'].name, float(item['balance'])])
    ws.append([_("إجمالي الأصول"), float(report['total_assets'])])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

    # Liabilities
    ws.append([])
    ws.append([_("الخصوم")])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color='FF0000')
    for item in report['liabilities']:
        ws.append([item['account'].name, float(item['balance'])])
    
    # Equity
    ws.append([])
    ws.append([_("حقوق الملكية")])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color='A52A2A')
    for item in report['equity']:
        ws.append([item['account'].name, float(item['balance'])])
    ws.append([_("صافي ربح الفترة"), float(report['net_profit'])])
    
    ws.append([_("إجمالي الخصوم وحقوق الملكية"), float(report['total_liabilities_equity'])])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=balance_sheet.xlsx'
    wb.save(response)
    return response
