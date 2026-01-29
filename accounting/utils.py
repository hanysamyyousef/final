import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from datetime import datetime

class ExcelExportUtil:
    """أداة لتصدير التقارير المحاسبية إلى ملفات Excel"""

    @staticmethod
    def export_report(report_name, headers, data, filename=None):
        """تصدير تقرير عام إلى Excel"""
        if not filename:
            filename = f"{report_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_name
        ws.sheet_view.rightToLeft = True  # دعم العربية

        # التنسيقات
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # إضافة العنوان الرئيسي
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=report_name)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = alignment
        ws.row_dimensions[1].height = 30

        # إضافة العناوين
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = thin_border

        # إضافة البيانات
        for row_num, row_data in enumerate(data, 3):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.alignment = alignment
                cell.border = thin_border
                
                # تنسيق الأرقام
                if isinstance(cell_value, (int, float)):
                    cell.number_format = '#,##0.00'

        # ضبط عرض الأعمدة
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # إنشاء الاستجابة
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
